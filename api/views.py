import json
from datetime import datetime, timedelta

from django.core.exceptions import ValidationError
from django.db.models import Count
from django.db.utils import IntegrityError
from django.http import HttpResponse, JsonResponse
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side

from orders.models import Customer, Order
from robots.models import Robot, RobotModel


@method_decorator(csrf_exempt, name='dispatch')
class RobotView(View):
    """Создание и получение всех роботов."""

    def get(self, request):

        robots = Robot.objects.all()
        robots_serialized_data = []

        for robot in robots:
            robots_serialized_data.append({
                'serial': robot.serial,
                'model': robot.model,
                'version': robot.version,
                'created': robot.created,
            })

        data = {
            'robots': robots_serialized_data,
        }
        return JsonResponse(data)

    def post(self, request):

        robot_data = json.load(request)

        try:
            robot = Robot()

            robot.model = robot_data.get('model')
            robot.version = robot_data.get('version')
            robot.created = robot_data.get('created')
            robot.serial = f'{robot.model}-{robot.version}'

            robot.full_clean()
            robot.save()

            # Для полного вывода информации при создании объекта
            robot_data['serial'] = robot.serial

        except IntegrityError as error:
            empty_value = str(error).rsplit('.')
            message = f'Не заполнено поле {empty_value[1]}'
            return JsonResponse({'message': message}, status=400)
        except ValidationError as error:
            message = f'Ошибка валидации: {error.messages}'
            return JsonResponse({'message': message}, status=400)
        except Exception as error:
            message = f'Ошибка: {error}'
            return JsonResponse({'message': message}, status=400)

        return JsonResponse(robot_data, status=201)


class TableCreatedRobots(View):
    """
    Выгрузка отчета по показателям производства роботов в Excel-таблицу.
    """

    # Период (количество дней) за который берутся данные в отчет
    report_days = 7

    # Заголовки колонок таблицы
    headers = ['Модель', 'Версия', 'Количество за неделю']

    # Параметры форматирования таблицы
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    alignment = Alignment(horizontal="center")

    def get(self, *args, **kwargs):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = \
            'attachment; filename="created_robots.xlsx"'

        wb = Workbook()

        # Выборка всех моделей роботов (соответствует количеству листов)
        models = RobotModel.objects.all()

        # Переименование первого листа под первую модель робота
        ws = wb.active
        ws.title = models[0].name

        # Создание листов под каждую модель (начиная со второй)
        for model in models[1:]:
            wb.create_sheet(f'{model.name}')

        # Вычисление дат для периода запроса данных о производстве
        date_to = datetime.today()
        date_from = date_to - timedelta(days=self.report_days)

        # Выборка данных о производстве роботов за определенный период
        robots = Robot.objects.filter(created__range=(date_from, date_to))

        for model in models:
            # Переключение на лист текущей модели роботов
            ws = wb[model.name]

            # Создание промежуточной таблицы (словаря) с данными
            # версии робота и их количестве
            current_robots_model = robots.filter(
                model=model.name
            ).values('version').annotate(count=Count('version'))

            # Заполнение заголовков таблицы
            for col, val in enumerate(self.headers, start=1):
                cell = ws.cell(row=1, column=col, value=val)
                cell.alignment = self.alignment
                cell.border = self.border
                # Ширина колонок форматируется под ширину заголовков таблицы
                ws.column_dimensions[cell.column_letter].width = len(val) + 2

            # Заполнение таблицы
            row = 2
            for robot in current_robots_model:
                lst = [model.name, robot.get('version'), robot.get('count')]
                for col, val in enumerate(lst, start=1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.alignment = self.alignment
                    cell.border = self.border
                row += 1

        wb.save(response)
        return response


@method_decorator(csrf_exempt, name='dispatch')
class OrderView(View):
    """Добавление заказа на робота."""

    def post(self, request):

        order_data = json.load(request)

        try:
            order = Order()

            customer = Customer.objects.get(id=order_data.get('customer'))
            order.customer = customer
            order.robot_serial = order_data.get('robot_serial')

            order.full_clean()

            # Проверка наличия готового робота на складе для текущего заказа
            serial = order.robot_serial
            robots = Robot.objects.filter(serial=serial, sold=False)

            # Если робот есть на складе добавляем его в заказ
            # и отмечаем выбранного робота в категорию "проданные"
            if robots:
                robot = robots[0]
                order.purchase = robot
                robot.sold = True
                robot.save()

            order.save()

        except IntegrityError as error:
            empty_value = str(error).rsplit('.')
            message = f'Не заполнено поле {empty_value[1]}'
            return JsonResponse({'message': message}, status=400)
        except ValidationError as error:
            message = f'Ошибка валидации: {error}'
            return JsonResponse({'message': message}, status=400)
        except Exception as error:
            message = f'Ошибка: {error}'
            return JsonResponse({'message': message}, status=400)

        return JsonResponse(order_data, status=201)
